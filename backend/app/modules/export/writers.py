"""Turns built sheets into bytes.

CSV is written with a UTF-8 BOM: without it Excel on Windows reads the file as
the system codepage and mangles any non-ASCII name.
"""
import csv
import io
import zipfile
from typing import Iterable, List, Tuple

#: Excel tab names cannot exceed this, and cannot contain : \\ / ? * [ ]
EXCEL_TAB_LIMIT = 31
_EXCEL_FORBIDDEN = str.maketrans({c: "-" for c in ":\\/?*[]"})

CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
ZIP_MEDIA_TYPE = "application/zip"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: Built sheet: (label, filename stem, headers, rows)
BuiltSheet = Tuple[str, str, List[str], List[list]]


def _csv_bytes(headers: List[str], rows: Iterable[list]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def write_csv_single(sheet: BuiltSheet) -> bytes:
    _label, _stem, headers, rows = sheet
    return _csv_bytes(headers, rows)


def write_csv_zip(sheets: List[BuiltSheet], readme: str) -> bytes:
    """One .csv per sheet plus a README carrying the same provenance as the
    Summary sheet, so a loose folder of CSVs still explains itself."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("README.txt", readme.encode("utf-8"))
        for sheet in sheets:
            _label, stem, headers, rows = sheet
            archive.writestr(f"{stem}.csv", _csv_bytes(headers, rows))
    return buffer.getvalue()


def _tab_name(label: str, used: set) -> str:
    name = label.translate(_EXCEL_FORBIDDEN)[:EXCEL_TAB_LIMIT] or "Sheet"
    if name not in used:
        used.add(name)
        return name
    # Excel refuses duplicate tab names; suffix until unique.
    for suffix in range(2, 100):
        tail = f" {suffix}"
        candidate = f"{name[:EXCEL_TAB_LIMIT - len(tail)]}{tail}"
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise ValueError(f"Cannot find a unique tab name for {label!r}")


#: Columns are sized from the header plus this many rows. Measuring every row of
#: a large sheet costs more than the tidier width is worth.
WIDTH_SAMPLE_ROWS = 50
MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 42


def _column_widths(headers: List[str], rows: List[list]) -> List[int]:
    widths = [min(max(len(str(value)) + 2, MIN_COL_WIDTH), MAX_COL_WIDTH) for value in headers]
    for row in rows[:WIDTH_SAMPLE_ROWS]:
        for index, value in enumerate(row):
            if index < len(widths) and value is not None:
                widths[index] = min(max(widths[index], len(str(value)) + 2), MAX_COL_WIDTH)
    return widths


def write_xlsx(sheets: List[BuiltSheet]) -> bytes:
    """One workbook, one tab per sheet.

    write_only mode keeps memory flat by streaming rows straight to the zip
    member instead of building a cell object graph. The catch: <cols> and
    <sheetViews> are serialised ahead of the row data, so column widths and
    freeze panes MUST be set before the first append or they are silently
    dropped. Everything that does not depend on the rows is therefore configured
    up front.
    """
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook(write_only=True)
    used_names = set()
    bold = Font(bold=True)

    for label, _stem, headers, rows in sheets:
        worksheet = workbook.create_sheet(_tab_name(label, used_names))

        for index, width in enumerate(_column_widths(headers, rows), start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width

        if headers:
            worksheet.freeze_panes = "A2"
            if rows:
                last_column = get_column_letter(len(headers))
                worksheet.auto_filter.ref = f"A1:{last_column}{len(rows) + 1}"

        header_cells = []
        for value in headers:
            cell = WriteOnlyCell(worksheet, value=value)
            cell.font = bold
            header_cells.append(cell)
        worksheet.append(header_cells)

        for row in rows:
            worksheet.append(row)

    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()
